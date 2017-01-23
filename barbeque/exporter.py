from datetime import date

import openpyxl
from django.core.exceptions import ImproperlyConfigured
from django.db.models.fields import FieldDoesNotExist
from django.http import HttpResponse
from django.utils.encoding import force_text
from django.utils.translation import ugettext_lazy as _

from .compat import UnicodeWriter


class Exporter(object):
    def __init__(self, export_type, short_description, fields, header):
        self.export_type = export_type
        self.short_description = short_description
        self.fields = fields
        self.header = header

    def get_filename(self, modeladmin, extension=None):
        return '%s_%s.%s' % (
            force_text(modeladmin.model._meta).replace('.', '_'),
            date.today().strftime('%Y-%m-%d'),
            extension or self.export_type
        )

    def get_header(self, queryset):
        base_obj = queryset[0]
        columns = []
        for field in self.fields:
            obj = base_obj
            parts = field.split('__')
            obj_path = parts[:-1]
            field_name = parts[-1]

            for part in obj_path:
                obj = getattr(obj, part)
            try:
                columns.append(force_text(
                    obj._meta.get_field_by_name(field_name)[0].verbose_name))
            except FieldDoesNotExist:
                columns.append(field)

        return columns

    def get_data(self, queryset):
        return (self.fields, queryset.order_by('pk').values(*self.fields))

    def export_as_csv(self, modeladmin, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=%s' % (
            self.get_filename(modeladmin, 'csv'))

        return self.write_csv(queryset, response)

    def write_csv(self, queryset, fobj):
        with UnicodeWriter(fobj) as writer:
            if self.header:
                writer.writerow([force_text(name) for name in self.get_header(queryset)])

            data_fields, data = self.get_data(queryset)
            for row in data:
                writer.writerow([
                    self.get_value(field, row.get(field, '')) for field in data_fields])

        return fobj

    def export_as_xlsx(self, modeladmin, request, queryset):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=%s' % self.get_filename(
            modeladmin, 'xlsx')
        return self.write_xlsx(queryset, response)

    def write_xlsx(self, queryset, fobj):
        workbook = openpyxl.Workbook()
        sheet = workbook.worksheets[0]

        start_row = 1
        if self.header:
            for j, name in enumerate(self.get_header(queryset), 1):
                sheet.cell('%s%s' % (
                    openpyxl.utils.get_column_letter(j), start_row)).value = name
            start_row += 1

        data_fields, data = self.get_data(queryset)

        column_widths = [len(i) for i in data_fields]

        for i, row in enumerate(data, start_row):
            for j, field in enumerate(data_fields, 1):
                value = self.get_value(field, row.get(field, ''))

                if len(value) + 2 > column_widths[j - 1]:
                    column_widths[j - 1] = len(value) + 2

                sheet.cell('%s%s' % (
                    openpyxl.utils.get_column_letter(j), i)
                ).value = value.replace('\r', '').replace('\n', ' ')

        for j, width in enumerate(column_widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(j)].width = width

        workbook.save(fobj)

        return fobj

    def get_value(self, field, value):
        return force_text(value)


def action_export_factory(
    export_type='csv', short_description=None, fields=None, header=True, cls=Exporter
):
    if export_type == 'csv':
        def export_csv_func(*args, **kwargs):
            exporter = cls(export_type, short_description, fields, header)
            return exporter.export_as_csv(*args, **kwargs)
        export_csv_func.short_description = short_description or _('Export as CSV')
        return export_csv_func
    elif export_type == 'xlsx':
        def export_xlsx_func(*args, **kwargs):
            exporter = cls(export_type, short_description, fields, header)
            return exporter.export_as_xlsx(*args, **kwargs)
        export_xlsx_func.short_description = short_description or _('Export as XLSX')
        return export_xlsx_func
    else:
        raise ImproperlyConfigured('Invalid export type, choose one of: csv, xlsx')
